import math

import cartopy.feature as cfeature
import geopandas as gpd
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
import xarray as xr
from cartopy import crs as ccrs
from matplotlib.ticker import MultipleLocator
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from my_config import (
    Vars,
    DirsLocal,
)

countries_raster = xr.open_dataset(DirsLocal.dir_file_country_raster_report)

c = sns.color_palette("tab10")
consistent_colors = dict(
    zip(
        ["USA", "IDN", "ITA", "CHN", "IND", "JPN", "NGA", "COD", "Other", "EGY"],
        [c[0], c[1], c[2], c[3], c[5], c[6], c[7], c[8], c[9], c[4]],
    )
)


def update_excel_results(
    data,
    sheet_name,
    message: str,
    file_name=DirsLocal.dir_file_excel_submission,
    col_offset=0,
    rw_offset=0,
    clear_sheet=False,
) -> tuple[int, int]:
    # Load the existing Excel file
    book = load_workbook(file_name)

    # Check if the sheet exists, if not, create it
    if sheet_name not in book.sheetnames:
        sheet = book.create_sheet(sheet_name)
    else:
        sheet = book[sheet_name]

    if clear_sheet:
        # Clear the contents of the sheet
        for row in sheet.iter_rows():
            for cell in row:
                cell = None

    # Write the new data to the sheet starting from cell A1
    r_idx, c_idx = 0, 0
    for r_idx, row in enumerate(
        dataframe_to_rows(data, index=False, header=True), rw_offset + 2
    ):
        for c_idx, value in enumerate(row, col_offset + 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    sheet.cell(row=rw_offset + 1, column=col_offset + 1, value=message)

    book.save(file_name)

    return r_idx, c_idx


def _summary_weight(data, yrs):
    return (
        (data.sel(year=yrs) * cos_lat)
        .mean(["latitude", "longitude"])
        .mean(dim="year")
        .compute()
    )


def _summary(data, yrs):
    return data.sel(year=yrs).sum(["latitude", "longitude"]).mean(dim="year").compute()


country_lc_grouping = pd.read_excel(
    DirsLocal.dir_file_lancet_country_info,
    header=1,
)

countries = gpd.read_file(DirsLocal.dir_file_country_polygons)
countries = countries.rename(columns={"ISO_3_CODE": "country"})

population_over_65 = xr.open_dataarray(DirsLocal.dir_pop_elderly_file)
population_infants = xr.open_dataarray(DirsLocal.dir_pop_infants_file)
population_over_75 = xr.open_dataarray(DirsLocal.dir_pop_above_75_file)

population_over_65["age_band_lower_bound"] = 65
population = xr.concat(
    [population_infants, population_over_65], dim="age_band_lower_bound"
)
population.name = "population"

heatwave_metrics_files = sorted(DirsLocal.dir_results_heatwaves.glob("*.nc"))
heatwave_metrics = xr.open_mfdataset(heatwave_metrics_files, combine="by_coords")

exposures_over65 = xr.open_dataset(
    DirsLocal.dir_results_pop_exposure
    / f"heatwave_exposure_change_over65_multi_threshold_{Vars.year_min_analysis}-{Vars.year_max_analysis}_worldpop.nc"
)

exposures_infants = xr.open_dataset(
    DirsLocal.dir_results_pop_exposure
    / f"heatwave_exposure_change_infants_multi_threshold_{Vars.year_min_analysis}-{Vars.year_max_analysis}_worldpop.nc"
)

exposures_over65 = exposures_over65.to_array()
exposures_over65["age_band_lower_bound"] = 65
exposures_infants = exposures_infants.to_array()
exposures_infants["age_band_lower_bound"] = 0
exposures_over65 = exposures_over65.squeeze().drop_vars("variable")
exposures_infants = exposures_infants.squeeze().drop_vars("variable")

exposures_change = xr.concat(
    [exposures_infants, exposures_over65],
    dim=pd.Index([0, 65], name="age_band_lower_bound"),
)
total_exposures = exposures_change.sum(["latitude", "longitude"])

total_exposures_change_over65 = total_exposures.sel(
    age_band_lower_bound=65, drop=True
).to_dataframe("elderly")
total_exposures_change_infants = total_exposures.sel(
    age_band_lower_bound=0, drop=True
).to_dataframe("infants")

# Load exposure absolute values (not exposure to change)
exposures_abs = xr.open_dataset(DirsLocal.dir_file_all_exposure_abs)
population_df = (
    population.sum(dim=["latitude", "longitude"]).to_dataframe().reset_index()
)
exposures_abs_df = (
    exposures_abs.sum(dim=["latitude", "longitude"]).to_dataframe().reset_index()
)
exposures_abs_df = exposures_abs_df.rename({"heatwave_days": "total heatwave days"})

with pd.ExcelWriter(
    DirsLocal.dir_results_pop_exposure / "indicator_1_1_2_heatwaves_summary.xlsx"
) as writer:
    pd.merge(population_df, exposures_abs_df).to_excel(
        writer, sheet_name="Global", index=False
    )

# Load the country exposure results
country_weighted = xr.open_dataset(
    DirsLocal.dir_worldpop_exposure_by_region
    / f"countries_heatwaves_exposure_weighted_change_{Vars.year_min_analysis}-{Vars.year_max_analysis}_worldpop.nc"
)
country_exposure_change = xr.open_dataset(
    DirsLocal.dir_worldpop_exposure_by_region
    / f"countries_heatwaves_exposure_change_{Vars.year_min_analysis}-{Vars.year_max_analysis}_worldpop.nc"
)
country_exposure_abs = xr.open_dataset(
    DirsLocal.dir_worldpop_exposure_by_region
    / f"countries_heatwaves_exposure_{Vars.year_min_analysis}-{Vars.year_max_analysis}_worldpop.nc"
)

country_exposure_abs_df = (
    country_exposure_abs.sel(year=slice(1980, None))
    .to_dataframe()
    .reset_index()
    .rename(columns={"country": "ISO3"})
)

# todo I needed to comment this out since the column names have changed
# country_exposure_abs_df = country_exposure_abs_df.drop(columns="exposures_weighted")
# country_exposure_abs_df = country_exposure_abs_df.rename(
#     columns={"exposures_total": "total heatwave days"}
# )

# fixme this file has repeated values for the same year
exposures_abs_lc_groups = xr.open_dataset(
    DirsLocal.dir_file_exposures_abs_by_lc_group_worldpop
)

# IMPORTANT need to use weighted average for HW 'raw' can't just do the sum across pixels cus that's bollocks.
cos_lat = np.cos(np.radians(heatwave_metrics.latitude))

hw_ref = _summary_weight(heatwave_metrics.heatwave_days, slice(1986, 2005))
hw_dec = _summary_weight(
    heatwave_metrics.heatwave_days, slice(2013, Vars.year_max_analysis)
)


# fixme the following code is not working since the rolling function cannot be calculated
# hw_rol = (
#     (heatwave_metrics.heatwave_days * cos_lat)
#     .mean(["latitude", "longitude"])
#     .rolling(year=10)
#     .mean()
#     .compute()
# )
#
# hw_rol.name = "heatwave_days"
#
# # hw_ref = _summary(heatwave_metrics.heatwave_days, slice(1986,2005))
# # hw_dec = _summary(heatwave_metrics.heatwave_days, slice(2013,max_year))
# # hw_rol = heatwave_metrics.heatwave_days.sum(['latitude', 'longitude']).rolling(year=10).mean().compute()
#
# po_ref = _summary(population, slice(1986, 2005))
# po_dec = _summary(population, slice(2013, max_year))
# po_rol = population.sum(["latitude", "longitude"]).rolling(year=10).mean().compute()
#
# ex_ref = _summary(exposures_abs.heatwave_days, slice(1986, 2005))
# ex_dec = _summary(exposures_abs.heatwave_days, slice(2013, max_year))
# ex_rol = (
#     exposures_abs.heatwave_days.sum(["latitude", "longitude"])
#     .rolling(year=10)
#     .mean()
#     .compute()
# )
# ex_rol.name = "heatwave_person_days"
# ex_rol.to_dataframe().dropna().to_csv(
#     results_folder / "heatwave_exposure_days_10_year_rolling_mean.csv"
# )
# hw_rol.to_dataframe().dropna().to_csv(
#     results_folder / "heatwave_days_10_year_rolling_mean.csv"
# )
# po_rol.to_dataframe().dropna().to_csv(
#     results_folder / "population_10_year_rolling_mean.csv"
# )
# # By LC group
# (100 * (po_rol - po_ref) / po_ref).to_dataframe().unstack(0).plot()
# plt.show()
#
# ax = (100 * (ex_rol - ex_ref) / ex_ref).to_dataframe("heatwave days").unstack(0).plot()
# ax.axhline(0)
# plt.show()


def plot_heatwaves_days(plot_data, slice_range=slice(1986, 2005)):
    # > Important when showing averages, don't do average of weighted number of days per country since you need to have
    # it always population wieghted, otherwise HW for china counts the same as HW for luxembourg. lc_map =
    # countries.dissolve('LC Grouping')

    # plot_data = plot_data.where(population.sel(age_band_lower_bound=65) > 10)
    plot_data = plot_data.sel(year=slice_range).mean(dim="year")
    f, ax = plt.subplots(
        figsize=(6, 3), subplot_kw=dict(projection=Vars.map_projection)
    )

    plot_data.heatwave_days.plot(vmax=15, transform=ccrs.PlateCarree(), ax=ax)
    plt.show()


def plot_change_in_heatwaves(year=Vars.year_max_analysis):
    plot_data = heatwave_metrics.sel(year=year) - heatwave_metrics.sel(
        year=slice(Vars.year_reference_start, Vars.year_reference_end)
    ).mean(dim="year")
    plot_data = plot_data.assign_coords(
        longitude=(((plot_data.longitude + 180) % 360) - 180)
    ).sortby("longitude", ascending=False)
    land_mask = countries_raster["OBJECTID"] < 2000
    plot_data = land_mask * plot_data
    # Define the structure of the figure with space for the colorbar
    f, ax = plt.subplots(
        1,
        1,
        figsize=(7, 5),
        subplot_kw=dict(projection=Vars.map_projection),
        constrained_layout=True,
    )

    cbar_ax = f.add_axes([0.1, 0.1, 0.8, 0.03])  # Position of the colorbar

    # plot_data_downsized = plot_data.heatwave_days.coarsen(
    #     latitude=10, longitude=10, boundary="trim"
    # ).mean()
    ax.coastlines(linewidths=0.25, resolution="110m")
    ax.add_feature(cfeature.OCEAN)

    # Plot the downsized data
    plot_data.heatwave_days.plot(
        transform=ccrs.PlateCarree(),
        ax=ax,
        vmin=-50,
        vmax=50,
        cmap="RdBu_r",
        cbar_kwargs={
            "label": "Change in number of heatwave days",
            "orientation": "horizontal",
            "cax": cbar_ax,
        },
    )
    for spine in ax.spines.values():
        spine.set_edgecolor("lightgray")
        spine.set_linewidth(0.5)
    ax.set_title(f"Change in {year} relative to 1986-2005 baseline")
    # ax.figure.savefig(dir_figures / f"map_hw_change_{year}.pdf")
    ax.figure.savefig(DirsLocal.dir_figures / f"map_hw_change_{year}.png", dpi=300)
    plt.show()


def plot_average_number_heatwaves_experienced():
    population = xr.concat(
        [
            population_infants,
            population_over_65,
            population_over_75.assign_coords(age_band_lower_bound=75),
        ],
        dim="age_band_lower_bound",
    )
    population.name = "population"
    norm_pop_exposure = exposures_abs.sum(["latitude", "longitude"]) / population.sel(
        year=slice(1980, Vars.year_max_analysis)
    ).sum(["latitude", "longitude"])
    norm_pop_exposure_df = norm_pop_exposure.to_dataframe()["heatwave_days"].unstack(1)
    norm_pop_exposure_df.to_csv(
        DirsLocal.dir_results_pop_exposure / "heatwave_days_experienced.csv"
    )
    norm_pop_exposure_df = norm_pop_exposure_df.reset_index()
    norm_pop_exposure_df = norm_pop_exposure_df.set_index("year")
    # norm_pop_exposure_df.columns = norm_pop_exposure_df.columns.droplevel(0)
    norm_pop_exposure_df.columns.name = "Age group"
    print("Average exposure per age group\n", norm_pop_exposure_df.tail(5).round(1))

    # I am excluding the over 75 since I do not have data for the whole period
    mean_reference_period = norm_pop_exposure_df.loc[
        (norm_pop_exposure_df.index >= Vars.year_reference_start)
        & (norm_pop_exposure_df.index <= Vars.year_reference_end),
        [0, 65],
    ].mean()
    print(
        "Percentage increment from reference period\n",
        round(
            (norm_pop_exposure_df[[0, 65]].tail(5) - mean_reference_period)
            / mean_reference_period
            * 100
        ),
    )

    norm_pop_exposure_df = norm_pop_exposure_df.rename(
        columns={0: "Infants", 65: "Over 65", 75: "Over 75"}
    )

    update_excel_results(
        clear_sheet=True,
        data=norm_pop_exposure_df.reset_index(),
        sheet_name=SheetsFinalSubmission.global_average,
        file_name=DirsLocal.dir_file_excel_submission,
        message="Average number of heatwaves days experienced per year by older adults (over 65 and over 75)  and infants",
    )

    f, ax = plt.subplots(figsize=(7, 4))
    norm_pop_exposure_df.plot(ax=ax)

    ylim = ax.get_ylim()
    ax.set(
        ylim=(0, math.ceil(ylim[1])),
        ylabel="Average heatwave days per year",
        xlabel="Year",
    )
    ax.yaxis.set_major_locator(MultipleLocator(5))
    ax.grid(ls="--", color="lightgray")
    ax.legend(frameon=False)
    sns.despine()
    plt.tight_layout()
    ax.figure.savefig(DirsLocal.dir_figures / "global_hw_per_person.pdf")
    plt.show()

    # - Don't really do it (according to Xiang isn't that obvious), just report % changes in HW, Persons,
    # and person-days between two reference periods - Choose a 'recent' period, could do ten-years to date so
    # 2013-2022, bit random. Otherwise 2010-2020
    exposures_abs_rolling = norm_pop_exposure_df.rolling(10).mean().dropna()
    exposures_abs_rolling.unstack().to_csv(
        DirsLocal.dir_results_pop_exposure
        / "heatwave_days_experienced_10_year_rolling_mean.csv"
    )


def plot_total_number_heatwaves_experienced():
    plot_data = exposures_abs.sum(["latitude", "longitude"])
    plot_data = plot_data.to_dataframe()["heatwave_days"].unstack(1)
    print("Total exposure billions", round(plot_data.iloc[-1] / 1e9, 1))

    plot_data = plot_data.T
    plot_data = plot_data.reset_index()
    plot_data = plot_data.set_index("year")
    # plot_data.columns = plot_data.columns.droplevel(0)
    plot_data = plot_data.rename(columns={0: "Infants", 65: "Over 65", 75: "Over 75"})
    plot_data.reset_index(inplace=True)

    _, col_max = update_excel_results(
        clear_sheet=True,
        data=plot_data,
        sheet_name=SheetsFinalSubmission.global_total,
        file_name=DirsLocal.dir_file_excel_submission,
        message="Total number of heatwaves days experienced per year by older adults (over 65 and over 75)  and infants",
    )

    # total values, I am combining over 65 and under 1 since over 75 is subset of over 65
    summary_ages = plot_data.set_index("year")
    summary_ages["total"] = (
        summary_ages["Over 65"] + summary_ages["Infants"]
    )  # Calculate the percentage increment
    # increase compared to the previous year
    percentage_increment = summary_ages.pct_change() * 100
    print(
        "Percentage increment by year - all groups\n",
        round(percentage_increment.tail(5)),
    )
    mean_reference_period = summary_ages.loc[
        (summary_ages.index >= Vars.year_reference_start)
        & (summary_ages.index <= Vars.year_reference_end)
    ].mean()
    print(
        "Percentage increment from reference period\n",
        round(
            (summary_ages.tail(5) - mean_reference_period) / mean_reference_period * 100
        ),
    )

    update_excel_results(
        data=percentage_increment,
        sheet_name=SheetsFinalSubmission.global_total,
        file_name=DirsLocal.dir_file_excel_submission,
        message="Percentage change (current year compared to year before) in total number of heatwaves days experienced per year by older adults (over 65 and over 75) and infants",
        col_offset=col_max + 1,
    )

    plot_data = plot_data.set_index("year")

    f, ax = plt.subplots(figsize=(7, 4))

    (plot_data / 1e9).plot(ylabel="Billion person-days", ax=ax)

    ax.set(ylim=(0, 16))
    ax.legend(frameon=False)
    sns.despine()
    ax.grid(ls="--", color="lightgray")
    ax.yaxis.set_major_locator(MultipleLocator(2))
    plt.tight_layout()
    ax.figure.savefig(DirsLocal.dir_figures / "heatwaves_exposure_total.pdf")
    plt.show()


def plot_country_exposure(slice_range=slice(1986, 2005)):
    plot_df = (
        country_exposure_abs.exposures_weighted.sel(age_band_lower_bound=65, drop=True)
        .sel(year=slice_range)
        .mean(dim="year")
        .to_dataframe()
        .reset_index()
    )
    plot_df = plot_df.merge(countries).set_geometry("geometry")

    plot_df.plot(column="exposures_weighted", vmin=0, vmax=20, legend=True)
    plt.tight_layout()
    plt.show()


def plot_country_change():
    ref = (
        country_exposure_abs.exposures_weighted.sel(age_band_lower_bound=65, drop=True)
        .sel(year=slice(Vars.year_reference_start, Vars.year_reference_end))
        .mean(dim="year")
        .to_dataframe()
    )
    yr = (
        country_exposure_abs.exposures_weighted.sel(age_band_lower_bound=65, drop=True)
        .sel(year=slice(2013, Vars.year_max_analysis))
        .mean(dim="year")
        .to_dataframe()
    )

    (
        (yr - ref)
        .reset_index()
        .merge(countries, on="country")
        .set_geometry("geometry")
        .plot(column="exposures_weighted", legend=True, vmin=0, vmax=14, cmap="plasma")
    )
    plt.show()


def plot_absolute_exposure_lc_group(year=Vars.year_max_analysis):
    data_year_max = exposures_abs_lc_groups.exposures_weighted.sel(year=year)
    df_data_year_max = data_year_max.to_dataframe()

    ax = (
        df_data_year_max.exposures_weighted.unstack(1)
        .rename_axis(index="", columns="Heatwave days")
        .rename(index={"South and Central America": "South and \nCentral America"})
        .plot.bar(
            ylabel="days/year",
            title=f"Heatwave days per vulnerable person in {year}",
        )
        .legend(
            bbox_to_anchor=(1.04, 0.5),
            loc="center left",
            borderaxespad=0,
            title="Age group",
        )
    )
    plt.tight_layout()
    ax.figure.savefig(DirsLocal.dir_figures / f"heatwave_days_lc_group_{year}.pdf")
    plt.show()


def plot_absolute_and_change_exposure_range_years_lc_group():
    yr = (
        exposures_abs_lc_groups.exposures_weighted
        # .sel(age_band_lower_bound=65, drop=True)
        .sel(year=slice(2013, Vars.year_max_analysis))
        .mean(dim="year")
        .to_dataframe()
    )

    ax = (
        yr.exposures_weighted.unstack(1)
        .rename_axis(index="", columns="Heatwave days")
        .rename(index={"South and Central America": "South and \nCentral America"})
        .plot.bar(
            ylabel="days/year",
            title=f"Heatwave days per vulnerable person\n 10 year mean 2013-{Vars.year_max_analysis}",
        )
        .legend(
            bbox_to_anchor=(1.04, 0.5),
            loc="center left",
            borderaxespad=0,
            title="Age group",
        )
    )
    plt.tight_layout()
    ax.figure.savefig(
        DirsLocal.dir_figures
        / f"heatwave_days_lc_group_2013-{Vars.year_max_analysis}.pdf"
    )
    plt.show()

    ref = (
        exposures_abs_lc_groups.exposures_weighted
        # .sel(age_band_lower_bound=65, drop=True)
        .sel(year=slice(Vars.year_reference_start, Vars.year_reference_end))
        .mean(dim="year")
        .to_dataframe()
    )

    e = (
        (yr - ref)
        .exposures_weighted.unstack(1)
        .rename_axis(index="", columns="Heatwave days")
    )
    ax = (
        e.rename(index={"South and Central America": "South and \nCentral America"})
        .plot.bar(
            ylabel="days/year",
            title=f"Mean change in heatwave days per vulnerable person by region\n "
            f"from {Vars.year_reference_start}-{Vars.year_reference_end} "
            f"to 2013-{Vars.year_max_analysis} ",
        )
        .legend(
            bbox_to_anchor=(1.04, 0.5),
            loc="center left",
            borderaxespad=0,
            title="Age group",
        )
    )
    plt.tight_layout()
    ax.figure.savefig(
        DirsLocal.dir_figures
        / f"heatwave_days_change_to_baseline_lc_group_2013-{Vars.year_max_analysis}.pdf"
    )
    plt.show()

    p = (
        (100 * (yr - ref) / ref)
        .exposures_weighted.unstack(1)
        .rename_axis(index="", columns="Heatwave days")
    )
    p.columns = ["Infants", "65+"]
    ax = (
        p.rename(index={"South and Central America": "South and \nCentral America"})
        .plot.bar(
            ylabel="%",
            title=f"Increase in mean heatwave days per by region\n in 2013-{Vars.year_max_analysis} relative to baseline",
        )
        .legend(
            bbox_to_anchor=(1.04, 0.5),
            loc="center left",
            borderaxespad=0,
            title="Age group",
        )
    )
    plt.tight_layout()
    ax.figure.savefig(
        DirsLocal.dir_figures
        / f"heatwave_days_pct_to_baseline_lc_group_2013-{Vars.year_max_analysis}.pdf"
    )
    plt.show()


def plot_exposure_vulnerable_to_change_heatwave():
    plot_data = (
        exposures_change.sum(["latitude", "longitude"])
        .to_dataframe("")
        .unstack("age_band_lower_bound")
    )
    plot_data.columns = ["infants", "over 65"]
    plot_data = plot_data[["over 65", "infants"]]

    f, ax = plt.subplots(constrained_layout=True)
    ax = plot_data.plot.bar(stacked=True, width=0.89, ax=ax)
    ax.set_ylabel("Billion person-days")
    ax.set_title(
        "Exposures of vulnerable populations to \nchange in heatwave occurance"
    )
    ax.legend(title="Age")

    # NOTE: wasn't an easy way to set the different hatches so have to set manually the indexes
    for p in ax.patches[:20]:
        p.set_hatch("...")
        p.set_edgecolor("C0")
        p.set_facecolor("w")

    for p in ax.patches[44:65]:
        p.set_hatch("xxxx")
        p.set_edgecolor("C1")
        p.set_facecolor("w")

    plt.savefig(
        DirsLocal.dir_figures
        / f"heatwave person-days hybrid w newborn 1980-{Vars.year_max_analysis}.pdf"
    )
    plt.show()


def plot_exposure_vulnerable_absolute_heatwave():
    # Absolute exposures
    plot_data = (
        exposures_abs.sum(["latitude", "longitude"])
        .to_dataframe()
        .unstack("age_band_lower_bound")
    )

    plot_data.columns = ["infants", "over 65", "over 75"]

    f, ax = plt.subplots(constrained_layout=True)

    ax = plot_data.plot.bar(stacked=True, width=0.89, ax=ax)
    ax.set_ylabel("Billion person-days")
    ax.set_title("Exposures of vulnerable populations to heatwaves")
    ax.legend(title="Age ")
    plt.show()


def plot_exposure_vulnerable_to_change_by_country_heatwave(age_band=65):
    country_lc_grouping = pd.read_excel(
        DirsLocal.dir_file_lancet_country_info,
        header=1,
    )
    list(country_lc_grouping[country_lc_grouping.ISO3 == "IDN"]["Country Name to use"])
    var = "heatwave_days"

    top_codes = (
        country_exposure_change[var]
        .sel(
            year=slice(2015, Vars.year_max_analysis),
            age_band_lower_bound=age_band,
            drop=True,
        )
        .mean(dim="year")
        .to_dataframe()
        .sort_values(by=var, ascending=False)
        .head(5)[var]
        .index.to_list()
    )
    selected_data_list = []

    # Loop through each country code
    for country_code in top_codes:
        # Select the data for the current country
        selected_data = country_exposure_change[var].sel(country=country_code)

        # Append the selected data to the list
        selected_data_list.append(selected_data)

    # You now have a list of xarray DataArrays, one for each country
    # You can combine these into a single DataArray or Dataset if needed
    combined_data = xr.concat(selected_data_list, dim="country")

    # Sort and show the top 5 for a given year
    # top_codes = (country_exposure[var]
    #              .sel(year=slice(2015,2020), age_band_lower_bound=age_band, drop=True)
    #              .mean(dim='year')
    #              .to_dataframe()
    #              .sort_values(by=var, ascending=False)
    #              .head(5)[var].index.to_list()
    #             )

    results = (
        combined_data.sel(age_band_lower_bound=age_band, drop=True)
        .to_dataframe()[var]
        .unstack()
        .T
    )

    total_exposures_over65 = (
        country_exposure_change.sel(age_band_lower_bound=age_band, drop=True)
        .sum(dim="country")
        .to_dataframe()
        .reset_index()
    )

    # Difference between sum of top5 countries and total gives the 'other' category
    results["Other"] = np.array(total_exposures_over65["heatwave_days"]) - np.array(
        results.sum(axis=1)
    )
    # invert column order
    results = results[results.columns[::-1]]

    f, ax = plt.subplots(figsize=(6.2, 2.5))
    (results / 1e9).plot.bar(stacked=True, width=0.9, ax=ax)

    ax.set(
        xlabel="Year",
        ylabel="Billion person-days",
        title=f"Exposures of age={age_band} to \nchange in heatwave occurance",
    )
    ax.xaxis.set_tick_params(labelsize="small")
    ax.yaxis.set_tick_params(labelsize="small")

    # Manually order the legend
    handles, labels = ax.get_legend_handles_labels()
    d = dict(zip(labels, handles))
    iso_codes = dict(zip(labels, handles)).keys()

    ordered_handles = [d[l] for l in iso_codes]
    ordered_labels = [
        (
            country_lc_grouping["Country Name to use"][
                country_lc_grouping.ISO3 == i
            ].iloc[0]
            if i != "Other"
            else "Other"
        )
        for i in iso_codes
    ]

    ordered_handles = [d[l] for l in iso_codes]
    ax.legend(ordered_handles, ordered_labels, fontsize="small")

    plt.tight_layout()
    f.savefig(
        DirsLocal.dir_figures
        / f"hw_exposure_age_{age_band}_countries_1980-{Vars.year_max_analysis}.pdf"
    )
    plt.show()


def plot_exposure_vulnerable_absolute_by_country_heatwave(age_band=0, max_year=2024):
    var = "heatwave_days"

    top_codes = (
        country_exposure_change[var]
        .sel(year=slice(2015, max_year), age_band_lower_bound=age_band, drop=True)
        .mean(dim="year")
        .to_dataframe()
        .sort_values(by=var, ascending=False)
        .head(5)[var]
        .index.to_list()
    )

    var = "exposures_total"

    _total_exposures = exposures_abs.sum(["latitude", "longitude"])
    _total_exposures = (
        _total_exposures.sel(age_band_lower_bound=age_band, drop=True)
        .to_dataframe()
        .heatwave_days
    )

    plot_data = (
        country_exposure_abs[var]
        .sel(country=country_exposure_abs.country.isin(top_codes))
        .sel(age_band_lower_bound=age_band, year=slice(1980, max_year), drop=True)
        .to_dataframe()[var]
        .unstack()
        .T
    )

    # Difference between sum of top5 countries and total gives the 'other' category
    plot_data["Other"] = _total_exposures - plot_data.sum(axis=1)
    # invert column order
    plot_data = plot_data[plot_data.columns[::-1]]

    f, ax = plt.subplots(figsize=(7, 4))
    (plot_data / 1e9).plot.bar(stacked=True, width=0.9, ax=ax, color=consistent_colors)

    title = "Exposures of individuals over 65 to heatwaves"
    if age_band == 0:
        title = "Exposures of infants to heatwaves"

    ax.set(
        xlabel="Year",
        ylabel="Billion person-days",
        title=title,
    )
    ax.yaxis.set_tick_params(labelsize="small")

    # Manually order the legend
    handles, labels = ax.get_legend_handles_labels()
    d = dict(zip(labels, handles))
    iso_codes = dict(zip(labels, handles)).keys()

    ordered_handles = [d[l] for l in iso_codes]
    ordered_labels = [
        (
            country_lc_grouping["Country Name to use"][
                country_lc_grouping.ISO3 == i
            ].iloc[0]
            if i != "Other"
            else "Other"
        )
        for i in iso_codes
    ]

    ordered_handles = [d[l] for l in iso_codes]
    ax.legend(ordered_handles, ordered_labels, frameon=False)

    ax.grid(True, ls="--", color="lightgray", axis="y")
    plt.tight_layout()
    sns.despine()
    f.savefig(
        DirsLocal.dir_figures
        / f"hw_exposure_age_{age_band}_countries_1980-{max_year}.pdf"
    )
    plt.show()

    # calculate exposures for each country
    var = "heatwave_days"
    all_country_codes = (
        country_exposure_change[var]
        .sel(year=slice(2015, max_year), age_band_lower_bound=age_band, drop=True)
        .mean(dim="year")
        .to_dataframe()
        .sort_values(by=var, ascending=False)[var]
        .index.to_list()
    )

    var = "exposures_total"
    data_export = (
        country_exposure_abs[var]
        .sel(country=country_exposure_abs.country.isin(all_country_codes))
        .sel(age_band_lower_bound=age_band, year=slice(1980, max_year), drop=True)
        .to_dataframe()[var]
        .reset_index()
    )

    if age_band == 0:
        sheet = SheetsFinalSubmission.country_infants
    elif age_band == 65:
        sheet = SheetsFinalSubmission.country_over65
    elif age_band == 75:
        sheet = SheetsFinalSubmission.country_over75
    else:
        raise ValueError("Invalid age band")
    update_excel_results(
        clear_sheet=True,
        data=data_export,
        sheet_name=sheet,
        file_name=DirsLocal.dir_file_excel_submission,
        message="Total heatwave person-days experienced by country",
    )


def plot_exposure_by_hdi():
    # Load aggregated by hdi and WHO region data
    hdi_exposure = xr.open_dataset(
        DirsLocal.dir_worldpop_exposure_by_region
        / f"hdi_regions_heatwaves_exposure_{Vars.year_min_analysis}"
        f"-{Vars.year_max_analysis}_worldpop.nc"
    )

    # hdi_exposure_change = xr.open_dataset(
    #     dir_worldpop_exposure_by_region
    #     / f"hdi_regions_heatwaves_exposure_change_{Vars.year_min_analysis}-{Vars.year_max_analysis}_worldpop.nc"
    # )

    hdi_exposure_df = hdi_exposure.to_dataframe().reset_index().dropna()
    hdi_exposure_df = hdi_exposure_df.rename(
        columns={
            "exposures_total": "total heatwave days",
            "level_of_human_development": "HDI class",
            "age_band_lower_bound": "Age group",
            "exposures_weighted": "Average heatwave days",
            "year": "Year",
        }
    )

    plot_data = hdi_exposure_df.drop(columns="total heatwave days")

    df_combined_ages = plot_data.groupby(["Year", "HDI class"]).mean()
    df_combined_ages = (
        df_combined_ages["Average heatwave days"].unstack(level=1).round(2).dropna()
    )
    print(df_combined_ages)

    _, col_max = update_excel_results(
        clear_sheet=True,
        data=hdi_exposure_df,
        sheet_name=SheetsFinalSubmission.hdi_group,
        file_name=DirsLocal.dir_file_excel_submission,
        message="Average number of heatwaves days experienced per year by older adults (over 65) and infants combined",
    )

    print(df_combined_ages.pct_change().round(3) * 100)

    update_excel_results(
        data=(df_combined_ages.pct_change().round(3) * 100).reset_index(),
        sheet_name=SheetsFinalSubmission.hdi_group,
        file_name=DirsLocal.dir_file_excel_submission,
        message="Percentage change from previous year (current year compared to year before) in average number of heatwaves days experienced per year by older adults (over 65) and infants combined",
        col_offset=col_max + 1,
    )

    plot_data = plot_data[plot_data["HDI class"] != ""]

    f, axs = plt.subplots(
        2, 1, constrained_layout=True, sharex=True, sharey=True, figsize=(7, 7)
    )

    for ix, age in enumerate(plot_data["Age group"].unique()):
        plot_data_age = plot_data[plot_data["Age group"] == age]
        sns.lineplot(
            data=plot_data_age,
            x="Year",
            y="Average heatwave days",
            hue="HDI class",
            ax=axs[ix],
        )
        title = "Infants" if age == 0 else "Over 65"
        axs[ix].set_title(title)
        axs[ix].grid(True, ls="--", color="lightgray")

    axs[0].set(ylabel="Average heatwaves days per year")
    axs[1].set(ylabel="Average heatwaves days per year")
    # if rolling:
    #     axs[0].set(
    #         ylabel=f"{rolling} year rolling mean of population-weighted heatwave days",
    #     )
    axs[0].legend(title="HDI category", frameon=False)
    axs[1].legend().remove()

    sns.despine()
    plt.savefig(DirsLocal.dir_figures / "heatwave_days_by_hdi.pdf")
    plt.show()


def plot_exposure_by_who():
    who_exposure = xr.open_dataset(
        DirsLocal.dir_worldpop_exposure_by_region
        / f"who_regions_heatwaves_exposure_{Vars.year_min_analysis}-{Vars.year_max_analysis}_worldpop.nc"
    )

    who_exposure_df = who_exposure.to_dataframe().reset_index().dropna()

    update_excel_results(
        clear_sheet=True,
        data=who_exposure_df,
        sheet_name=SheetsFinalSubmission.who_region,
        file_name=DirsLocal.dir_file_excel_submission,
        message="Average heatwave person-days experienced by WHO region",
    )

    # plot the data
    plot_data = who_exposure_df.drop(columns=["exposures_total"]).rename(
        columns={
            "age_band_lower_bound": "Age group",
            "exposures_weighted": "Average heatwave days",
            "who_region": "WHO region",
        }
    )
    plot_data["WHO region"] = plot_data["WHO region"].replace(
        {
            "EMRO": "Eastern Med.",
            "EURO": "Europe",
            "AFRO": "Africa",
            "SEARO": "South-East Asia",
            "WPRO": "Western Pacific",
            "AMRO": "Americas",
        },
    )

    f, axs = plt.subplots(
        2, 1, constrained_layout=True, sharex=True, sharey=True, figsize=(7, 7)
    )

    for ix, age in enumerate(plot_data["Age group"].unique()):
        plot_data_age = plot_data[plot_data["Age group"] == age]
        sns.lineplot(
            data=plot_data_age,
            x="year",
            y="Average heatwave days",
            hue="WHO region",
            ax=axs[ix],
        )
        title = "Infants" if age == 0 else "Over 65"
        axs[ix].set_title(title)
        axs[ix].grid(True, ls="--", color="lightgray")

    axs[0].set(ylabel="Average heatwaves days per year")
    axs[1].set(ylabel="Average heatwaves days per year", xlabel="Year")

    axs[0].legend(title="WHO category", frameon=False, ncol=2)
    axs[1].legend().remove()

    sns.despine()
    plt.savefig(DirsLocal.dir_figures / "heatwave_days_by_who.pdf")
    plt.show()


def save_lc_data_to_excel():
    lc_exposures_abs_lc_groups_df = (
        exposures_abs_lc_groups.to_dataframe().reset_index().dropna()
    )
    lc_exposures_abs_lc_groups_df = lc_exposures_abs_lc_groups_df.rename(
        columns={
            "lc_group": "Lancet Countdown Region",
            "exposures_total": "total heatwave days",
        }
    )

    _, col_max = update_excel_results(
        clear_sheet=True,
        data=lc_exposures_abs_lc_groups_df,
        sheet_name=SheetsFinalSubmission.lc_region,
        file_name=DirsLocal.dir_file_excel_submission,
        message="Results by LC region",
    )


if __name__ == "__main__":
    pass


if __name__ == "__plot__":
    plot_heatwaves_days(plot_data=heatwave_metrics, slice_range=slice(1986, 2005))
    # world plot, slow to generate
    plot_heatwaves_days(
        plot_data=heatwave_metrics,
        slice_range=slice(2013, Vars.year_max_analysis),
    )
    plot_change_in_heatwaves(year=Vars.year_max_analysis)

    # plot trends
    plot_total_number_heatwaves_experienced()
    plot_average_number_heatwaves_experienced()
    plot_country_exposure(slice_range=slice(2013, Vars.year_max_analysis))
    plot_country_change()

    plot_absolute_exposure_lc_group(year=Vars.year_max_analysis)

    plot_absolute_and_change_exposure_range_years_lc_group()
    plot_exposure_vulnerable_to_change_heatwave()
    plot_exposure_vulnerable_absolute_heatwave()
